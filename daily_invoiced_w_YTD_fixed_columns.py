
# -*- coding: utf-8 -*-
import pandas as pd
import pyodbc
import yagmail
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Setup
dsn_name = "QuickBooks Data"
excel_file = f"Daily_Sales_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
recipient_email = ["steve@lenexamc.com", "andrew@lenexamc.com", "thomas@lenexamc.com", "mdefonso@gmail.com", "amy@lenexamc.com", "maryann@lenexamc.com", "dave@lenexamc.com"]
subject = "Sales Report, Invoiced Daily & YTD"
body = "Attached is sales (what was invoiced) for daily, yesterday, and year-to-date."
gmail_user = "stephenbastasch@gmail.com"
gmail_password = "nkvomogdhrswjyyr"  # Gmail app password

# Date ranges
today = datetime.today()
yesterday = today - timedelta(days=1)
ytd_start = datetime(today.year, 1, 1)
pytd_start = datetime(today.year - 1, 1, 1)
pytd_today = datetime(today.year - 1, today.month, today.day)

# SQL date format
fmt = "%Y-%m-%d"
date_ranges = {
    "today": today.strftime(fmt),
    "yesterday": yesterday.strftime(fmt),
    "ytd_start": ytd_start.strftime(fmt),
    "pytd_start": pytd_start.strftime(fmt),
    "pytd_today": pytd_today.strftime(fmt)
}

# Query
query = (
    f"SELECT I.TxnDate, IL.InvoiceLineAmount "
    f"FROM Invoice I INNER JOIN InvoiceLine IL ON I.TxnID = IL.TxnID "
    f"WHERE I.TxnDate IN ({{d '{date_ranges['today']}'}}, {{d '{date_ranges['yesterday']}'}}) "
    f"OR (I.TxnDate BETWEEN {{d '{date_ranges['ytd_start']}'}} AND {{d '{date_ranges['today']}'}}) "
    f"OR (I.TxnDate BETWEEN {{d '{date_ranges['pytd_start']}'}} AND {{d '{date_ranges['pytd_today']}'}})"
)

try:
    conn_str = f"DSN={dsn_name};"
    conn = pyodbc.connect(conn_str, autocommit=True)
    print("✅ Connected to QuickBooks via QODBC.")

    df = pd.read_sql(query, conn)
    df["TxnDate"] = pd.to_datetime(df["TxnDate"])

    def label_range(x):
        d = x.date()
        if d == today.date():
            return "Today"
        elif d == yesterday.date():
            return "Yesterday"
        elif ytd_start.date() <= d <= today.date():
            return "YTD"
        elif pytd_start.date() <= d <= pytd_today.date():
            return "PYTD"
        return None

    df["Range"] = df["TxnDate"].apply(label_range)
    totals = df.groupby("Range")["InvoiceLineAmount"].sum()

    today_total = totals.get("Today", 0)
    yesterday_total = totals.get("Yesterday", 0)
    ytd = totals.get("YTD", 0)
    pytd = totals.get("PYTD", 0)

    summary = pd.DataFrame({
        "Today": [today_total],
        "Yesterday": [yesterday_total],
        "Δ Today": [today_total - yesterday_total],
        "YTD": [ytd],
        "PYTD": [pytd],
        "Δ YTD": [ytd - pytd],
        "% Δ YTD": [((ytd - pytd) / pytd) if pytd != 0 else None]
    })

    print("\n📊 Invoiced Summary:")
    print(summary)

    summary.to_excel(excel_file, index=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        for col in range(1, 6):
            ws.cell(row=row, column=col).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=row, column=6).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=row, column=7).number_format = "0.0%"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    def autofit_columns(ws):
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    autofit_columns(ws)

    wb.save(excel_file)
    print(f"✅ Excel formatted, columns adjusted, and saved: {excel_file}")

    yag = yagmail.SMTP(gmail_user, gmail_password)
    yag.send(to=recipient_email, subject=subject, contents=body, attachments=excel_file)
    print("📧 Email sent.")

except Exception as e:
    print("❌ ERROR:")
    print(e)

finally:
    try:
        conn.close()
    except Exception as e:
        print(f"⚠️ Could not close connection: {e}")

    try:
        if os.path.exists(excel_file):
            os.remove(excel_file)
            print("🧹 Cleaned up: Excel file deleted.")
    except Exception as e:
        print(f"⚠️ Could not delete Excel file: {e}")
